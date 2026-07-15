"""
Conservative delete flow for Feishu search candidates.

The user-facing operation is "delete", but files are moved into a timestamped
backup directory and the workbook is backed up before rows are removed.
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import search
from multi_agent_pipeline import symbolic_root
from scripts.store_unindexed_questions import backup_workbook, ensure_workbook


@dataclass
class WorkbookDeleteTarget:
    bank: str
    workbook: Path
    rel_paths: list[str]
    rows: list[int]


@dataclass
class DeletePlan:
    rank: int
    chapter: str
    question_path: Path
    question_rel_path: str
    answer_paths: list[Path]
    workbook_targets: list[WorkbookDeleteTarget]
    candidate: dict[str, Any] = field(default_factory=dict)


@dataclass
class DeleteApplyResult:
    plan: DeletePlan
    dry_run: bool
    backup_dir: Path | None = None
    workbook_backups: list[Path] = field(default_factory=list)
    moved_files: list[Path] = field(default_factory=list)
    deleted_rows: int = 0


class FeishuDeleteService:
    def __init__(
        self,
        *,
        root: Path | None = None,
        symbolic: Path | None = None,
        dry_run: bool = False,
    ) -> None:
        self.root = Path(root or search.ROOT)
        self.symbolic = Path(symbolic or symbolic_root(self.root))
        self.dry_run = dry_run

    def prepare_plan(self, candidate: dict[str, Any], rank: int, chapter: str | None) -> DeletePlan:
        if rank < 1:
            raise ValueError("删除序号必须从 1 开始")
        if not chapter:
            raise ValueError("缺少章节信息，不能安全删除")

        raw_path = candidate.get("path")
        if not raw_path:
            raise ValueError("候选结果缺少题目路径")

        question_path, resolved_rel = self._resolve_question(raw_path, chapter)
        rel_candidates = candidate_rel_candidates(candidate, resolved_rel, self.root)
        question_rel_path = resolved_rel
        if rel_candidates:
            question_rel_path = rel_candidates[0]

        answer_paths = find_answer_files_for_question(question_path)
        workbook_targets = self._find_workbook_targets(chapter, rel_candidates)
        if not workbook_targets:
            raise ValueError("没有在主库或字母库 Excel 中找到这道题的记录，已取消删除")

        return DeletePlan(
            rank=rank,
            chapter=chapter,
            question_path=question_path,
            question_rel_path=question_rel_path,
            answer_paths=answer_paths,
            workbook_targets=workbook_targets,
            candidate=dict(candidate),
        )

    def apply_plan(self, plan: DeletePlan) -> DeleteApplyResult:
        if self.dry_run:
            return DeleteApplyResult(plan=plan, dry_run=True)

        backup_dir = Path(__file__).resolve().parents[1] / "backups" / f"delete_question_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        workbook_backups = self._backup_workbooks(plan, backup_dir)
        moved_files = self._move_files(plan, backup_dir)
        deleted_rows = self._delete_workbook_rows(plan)
        return DeleteApplyResult(
            plan=plan,
            dry_run=False,
            backup_dir=backup_dir,
            workbook_backups=workbook_backups,
            moved_files=moved_files,
            deleted_rows=deleted_rows,
        )

    def _find_workbook_targets(self, chapter: str, rel_candidates: list[str]) -> list[WorkbookDeleteTarget]:
        targets: list[WorkbookDeleteTarget] = []
        rel_set = {normalize_rel_path(rel) for rel in rel_candidates if rel}
        for bank, root in (("main", self.root), ("symbolic", self.symbolic)):
            workbook = resolve_workbook(root, chapter)
            if not workbook:
                continue
            wb = load_workbook(workbook)
            ws = wb.active
            headers = {str(cell.value).strip(): index for index, cell in enumerate(ws[1], 1) if cell.value}
            question_col = headers.get("题目名称")
            if not question_col:
                wb.close()
                continue
            rows: list[int] = []
            matched_rels: list[str] = []
            for row_index in range(2, ws.max_row + 1):
                rel = normalize_rel_path(ws.cell(row_index, question_col).value)
                if rel in rel_set:
                    rows.append(row_index)
                    if rel not in matched_rels:
                        matched_rels.append(rel)
            wb.close()
            if rows:
                targets.append(WorkbookDeleteTarget(bank=bank, workbook=workbook, rel_paths=matched_rels, rows=rows))
        return targets

    def _resolve_question(self, raw_path: object, chapter: str) -> tuple[Path, str]:
        try:
            if self.root.resolve() == Path(search.ROOT).resolve():
                path, rel, _ = search.resolve_question_path(
                    raw_path,
                    chapter_name=chapter,
                    update_excel=False,
                )
                return path, rel
        except OSError:
            pass

        path = Path(str(raw_path))
        if path.is_absolute():
            question_path = path
        elif path.exists():
            question_path = path.resolve()
        else:
            question_path = self.root / normalize_rel_path(path)
        return question_path, normalize_rel_path(relative_to_root(question_path, self.root))

    def _backup_workbooks(self, plan: DeletePlan, backup_dir: Path) -> list[Path]:
        backups: list[Path] = []
        for target in plan.workbook_targets:
            backup = backup_workbook(target.workbook, backup_dir, target.bank)
            if backup:
                backups.append(backup)
        return backups

    def _delete_workbook_rows(self, plan: DeletePlan) -> int:
        deleted = 0
        for target in plan.workbook_targets:
            wb, ws, _headers = ensure_workbook(target.workbook)
            for row_index in sorted(target.rows, reverse=True):
                ws.delete_rows(row_index, 1)
                deleted += 1
            wb.save(target.workbook)
            wb.close()
        return deleted

    def _move_files(self, plan: DeletePlan, backup_dir: Path) -> list[Path]:
        moved: list[Path] = []
        moved_pairs: list[tuple[Path, Path]] = []
        try:
            for source in unique_existing_paths([plan.question_path, *plan.answer_paths]):
                rel = relative_to_root(source, self.root)
                target = backup_dir / "files" / rel
                target.parent.mkdir(parents=True, exist_ok=True)
                if target.exists():
                    target = target.with_name(f"{target.stem}_{int(datetime.now().timestamp())}{target.suffix}")
                shutil.move(str(source), str(target))
                moved.append(target)
                moved_pairs.append((source, target))
        except Exception:
            for source, target in reversed(moved_pairs):
                if target.exists() and not source.exists():
                    source.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(target), str(source))
            raise
        return moved


def resolve_workbook(root: Path, chapter: str) -> Path | None:
    workbook = root / f"{chapter}.xlsx"
    if workbook.exists():
        return workbook
    matches = list(root.glob(f"*{chapter}*.xlsx"))
    return matches[0] if matches else None


def candidate_rel_candidates(candidate: dict[str, Any], resolved_rel: str, root: Path) -> list[str]:
    values: list[str] = []
    for key in ("name", "题目名称"):
        raw = candidate.get(key)
        if raw:
            values.append(normalize_rel_path(raw))
    if resolved_rel:
        values.append(normalize_rel_path(resolved_rel))
    raw_path = candidate.get("path")
    if raw_path:
        try:
            values.append(normalize_rel_path(Path(raw_path).resolve().relative_to(root.resolve())))
        except (OSError, ValueError):
            values.append(normalize_rel_path(search._rel_path_from_question_path(raw_path)))
    return dedupe_keep_order([value for value in values if value])


def normalize_rel_path(value: object) -> str:
    text = str(value or "").replace("\\", "/").strip()
    text = re.sub(r"^[A-Za-z]:/*", "", text)
    return text.lstrip("/")


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result


def unique_existing_paths(paths: list[Path]) -> list[Path]:
    seen: set[Path] = set()
    result: list[Path] = []
    for path in paths:
        resolved = path.resolve()
        if resolved in seen or not path.exists():
            continue
        result.append(path)
        seen.add(resolved)
    return result


def find_answer_files_for_question(question_path: Path) -> list[Path]:
    parts = question_path.parts
    stem = question_path.stem
    question_dir_index = None
    for index in range(len(parts) - 1, -1, -1):
        if parts[index].startswith("题目"):
            question_dir_index = index
            break
    if question_dir_index is None:
        return []

    answer_dir = Path(*parts[:question_dir_index]) / "答案"
    if not answer_dir.is_dir():
        return []

    found: list[Path] = []
    for ext in [".jpg", ".jpeg", ".png"]:
        for suffix in ["", "+", "++", "+++"]:
            path = answer_dir / f"{stem}{suffix}{ext}"
            if path.is_file():
                found.append(path)
    return found


def relative_to_root(path: Path, root: Path) -> Path:
    try:
        return path.resolve().relative_to(root.resolve())
    except ValueError:
        return Path(path.name)


def parse_delete_choice(text: str) -> int | None:
    match = re.fullmatch(r"-(\d+)", text.strip())
    if not match:
        return None
    return int(match.group(1))


def format_delete_confirmation(plan: DeletePlan) -> str:
    answer_names = "、".join(path.name for path in plan.answer_paths) or "未找到答案文件"
    row_count = sum(len(target.rows) for target in plan.workbook_targets)
    banks = "、".join(target.bank for target in plan.workbook_targets)
    return "\n".join([
        f"准备删除第 {plan.rank} 个候选：",
        f"章节：{plan.chapter}",
        f"题目：{plan.question_rel_path}",
        f"答案：{answer_names}",
        f"Excel：{banks}，{row_count} 行",
        "",
        "将先备份 Excel，并把题目/答案移动到 backups。",
        "回复 1 确认删除，回复 0 取消。",
    ])


def format_delete_success(result: DeleteApplyResult) -> str:
    if result.dry_run:
        return f"[dry-run] 已删除第 {result.plan.rank} 个候选。"
    return "\n".join([
        f"已删除第 {result.plan.rank} 个候选。",
        f"Excel 删除：{result.deleted_rows} 行",
        f"文件移动：{len(result.moved_files)} 个",
        f"备份：{result.backup_dir}",
    ])
