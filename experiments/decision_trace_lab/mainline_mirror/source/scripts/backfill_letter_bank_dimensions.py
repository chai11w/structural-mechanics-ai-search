"""Backfill the letter bank (帮做_字母库) with canonical long×width dimensions.

Reads a values file of ``{path, long_width}`` human-verified verdicts, backs up
the letter-bank xlsx files, then adds a ``长×宽`` column to every chapter sheet
(empty for rows without a verdict yet) and fills the matching rows. Existing
readers access columns by header (``题目名称`` / ``荷载`` / ``结构类型``), so
adding a new column is non-breaking.

Only the letter bank is touched; the main bank is intentionally left alone.
"""

from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

import openpyxl

DEFAULT_BANK_ROOT = Path("D:/桌面/答疑、帮做/结构力学/帮做_字母库")
DEFAULT_VALUES = (
    Path(__file__).resolve().parent.parent
    / "experiments" / "structure_dimension_eval" / "human_verdicts.json"
)
COLUMN_NAME = "长×宽"


def load_verdicts(path: Path) -> dict[str, str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    raw = data.get("verdicts")
    if not isinstance(raw, list):
        raise ValueError("values file must contain a verdicts list")
    verdicts: dict[str, str] = {}
    for item in raw:
        image_path = str(item.get("path") or "").replace("\\", "/").strip()
        long_width = str(item.get("long_width") or "").strip()
        if not image_path or not long_width:
            raise ValueError(f"verdict needs both path and long_width: {item!r}")
        verdicts[image_path] = long_width
    return verdicts


def backup_bank(root: Path) -> Path:
    backup_dir = root.parent / f"{root.name}_备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    backup_dir.mkdir(parents=True, exist_ok=False)
    for xlsx in root.glob("*.xlsx"):
        shutil.copy2(xlsx, backup_dir / xlsx.name)
    return backup_dir


def ensure_dimension_column(ws: Any, header_row: int = 1) -> tuple[int, bool]:
    """Return ``(column_index, added)`` for the dimension column, adding it if missing."""

    headers = [str(ws.cell(row=header_row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    if COLUMN_NAME in headers:
        return headers.index(COLUMN_NAME) + 1, False
    column = ws.max_column + 1
    ws.cell(row=header_row, column=column).value = COLUMN_NAME
    return column, True


def backfill_file(
    xlsx_path: Path, verdicts: Mapping[str, str], *, dry_run: bool
) -> tuple[int, set[str]]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    column, added = ensure_dimension_column(ws)
    matched: set[str] = set()
    for row in range(2, ws.max_row + 1):
        path = str(ws.cell(row=row, column=1).value or "").strip()
        if path not in verdicts:
            continue
        if not dry_run:
            ws.cell(row=row, column=column).value = verdicts[path]
        matched.add(path)
    if (matched or added) and not dry_run:
        wb.save(xlsx_path)
    wb.close()
    return len(matched), matched


def main() -> int:
    parser = argparse.ArgumentParser(description="Backfill letter-bank long×width dimensions.")
    parser.add_argument("--bank-root", type=Path, default=DEFAULT_BANK_ROOT)
    parser.add_argument("--values", type=Path, default=DEFAULT_VALUES)
    parser.add_argument("--dry-run", action="store_true", help="report matches without writing or backing up")
    args = parser.parse_args()

    root = args.bank_root.resolve()
    if not root.is_dir():
        raise SystemExit(f"letter bank root missing: {root}")
    verdicts = load_verdicts(args.values.resolve())

    total = 0
    found: set[str] = set()
    backup_dir: Path | None = None
    if not args.dry_run:
        backup_dir = backup_bank(root)
        print(f"backup={backup_dir}")
    for xlsx in sorted(root.glob("*.xlsx")):
        matched_count, matched_paths = backfill_file(xlsx, verdicts, dry_run=args.dry_run)
        if matched_count:
            total += matched_count
            found.update(matched_paths)
            print(f"{xlsx.name}: filled {matched_count}")
    print(f"filled_total={total} expected={len(verdicts)}")
    missing = [path for path in verdicts if path not in found]
    if missing:
        print("NOT_FOUND_IN_BANK:")
        for path in missing:
            print(f"  {path}")
    return 0 if not missing else 2


if __name__ == "__main__":
    raise SystemExit(main())
