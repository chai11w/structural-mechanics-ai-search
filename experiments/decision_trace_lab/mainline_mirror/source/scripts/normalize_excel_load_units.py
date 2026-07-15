"""
Normalize existing Excel load records to the unitless load-label rule.

Default mode is dry-run. Pass --apply to back up and rewrite live Excel files.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

import search  # noqa: E402
from scripts.audit_unindexed_questions import CHAPTERS  # noqa: E402


KNOWN_CORRECTIONS = {
    "5位移法/题目/3.jpg": [{"type": "均布", "raw": "5"}],
}


def symbolic_root(root: Path) -> Path:
    return root.parent / f"{root.name}_字母库"


def sheet_headers(ws) -> dict[str, int]:
    headers = {}
    for index, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = index
    return headers


def parse_loads(raw: Any) -> list[dict[str, Any]]:
    if not isinstance(raw, str) or not raw.strip():
        return []
    fixed = raw.replace("，", ",").replace("：", ":").replace("“", '"').replace("”", '"')
    try:
        parsed = json.loads(fixed)
    except json.JSONDecodeError:
        return []
    loads = parsed.get("loads", [])
    return loads if isinstance(loads, list) else []


def normalize_rel(raw: Any) -> str:
    return str(raw or "").replace("\\", "/").strip()


def normalize_loads(rel_path: str, loads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rel_path = normalize_rel(rel_path)
    if rel_path in KNOWN_CORRECTIONS:
        return [dict(item) for item in KNOWN_CORRECTIONS[rel_path]]

    normalized = []
    for item in loads:
        if not isinstance(item, dict):
            continue
        typ = search.normalize_load_type(item.get("type", ""), item.get("raw", ""))
        raw = search.strip_load_unit(item.get("raw", ""))
        record = {"type": typ, "raw": raw}
        if "original_raw" in item:
            record["original_raw"] = search.strip_load_unit(item.get("original_raw", ""))
        normalized.append(record)
    return normalized


def loads_json(loads: list[dict[str, Any]]) -> str:
    return json.dumps({"loads": loads}, ensure_ascii=False)


def backup_once(workbook: Path, backup_dir: Path, seen: set[Path]) -> None:
    if workbook in seen:
        return
    target = backup_dir / workbook.parent.name / workbook.name
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook, target)
    seen.add(workbook)


def normalize_workbook(workbook: Path, *, dry_run: bool, backup_dir: Path) -> dict[str, Any]:
    summary = {
        "workbook": str(workbook),
        "rows": 0,
        "changed": 0,
        "parse_errors": 0,
        "known_corrections": 0,
    }
    if not workbook.exists():
        return summary

    wb = load_workbook(workbook)
    ws = wb.active
    headers = sheet_headers(ws)
    if "题目名称" not in headers or "荷载" not in headers:
        wb.close()
        raise ValueError(f"{workbook}: missing required columns 题目名称/荷载")

    question_col = headers["题目名称"]
    loads_col = headers["荷载"]
    changed_rows = []

    for row_index in range(2, ws.max_row + 1):
        rel_path = normalize_rel(ws.cell(row_index, question_col).value)
        raw_loads = ws.cell(row_index, loads_col).value
        if not rel_path:
            continue
        summary["rows"] += 1
        loads = parse_loads(raw_loads)
        if raw_loads and not loads:
            summary["parse_errors"] += 1
            continue
        normalized = normalize_loads(rel_path, loads)
        if rel_path in KNOWN_CORRECTIONS:
            summary["known_corrections"] += 1
        new_value = loads_json(normalized)
        if str(raw_loads or "") != new_value:
            summary["changed"] += 1
            changed_rows.append((row_index, new_value))

    if changed_rows and not dry_run:
        backup_once(workbook, backup_dir, set())
        for row_index, value in changed_rows:
            ws.cell(row_index, loads_col).value = value
        wb.save(workbook)
    wb.close()
    return summary


def iter_workbooks(root: Path, symbolic: Path) -> list[Path]:
    paths = []
    for bank_root in (root, symbolic):
        for chapter in CHAPTERS:
            path = bank_root / f"{chapter}.xlsx"
            if path.exists():
                paths.append(path)
    return paths


def main() -> int:
    parser = argparse.ArgumentParser(description="Normalize Excel load records to unitless raw labels.")
    parser.add_argument("--root", help="main question bank root")
    parser.add_argument("--symbolic-root", help="symbolic question bank root")
    parser.add_argument("--apply", action="store_true", help="rewrite Excel files after backup")
    args = parser.parse_args()

    root = Path(args.root) if args.root else search.ROOT
    symbolic = Path(args.symbolic_root) if args.symbolic_root else symbolic_root(root)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = BASE / "backups" / f"normalize_load_units_{stamp}"

    summaries = [
        normalize_workbook(path, dry_run=not args.apply, backup_dir=backup_dir)
        for path in iter_workbooks(root, symbolic)
    ]

    total_rows = sum(item["rows"] for item in summaries)
    total_changed = sum(item["changed"] for item in summaries)
    total_errors = sum(item["parse_errors"] for item in summaries)
    total_known = sum(item["known_corrections"] for item in summaries)

    print(f"mode={'apply' if args.apply else 'dry-run'}")
    print(f"root={root}")
    print(f"symbolic_root={symbolic}")
    print(f"rows={total_rows} changed={total_changed} parse_errors={total_errors} known_corrections={total_known}")
    if args.apply and total_changed:
        print(f"backup_dir={backup_dir}")
    for item in summaries:
        if item["changed"] or item["parse_errors"]:
            print(
                f"{Path(item['workbook']).name}: rows={item['rows']} "
                f"changed={item['changed']} parse_errors={item['parse_errors']} "
                f"known_corrections={item['known_corrections']}"
            )
    return 0 if total_errors == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
