from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from PIL import Image

from scripts.feishu_store_flow import (
    FeishuStoreService,
    append_excel_record,
    classify_store_dimensions,
    format_store_confirmation,
)
from scripts.feishu_tiku_bot import build_parser


class _FakeQwen:
    def __init__(self):
        self.dimension_calls = 0
        self.image_paths = []

    def classify_image(self, image_path):
        self.image_paths.append(Path(image_path))
        return {
            "loads": [{"type": "均布", "raw": "q"}],
            "chapter_hint": "2静定结构",
            "chapter_confidence": 1.0,
            "chapter_evidence": "‘求图示桁架杆件轴力’",
        }

    def classify_structure_type(self, image_path):
        self.image_paths.append(Path(image_path))
        return {"structure_type": "桁架"}

    def recognize_dimensions(self, image_path, known_structure_type):
        self.image_paths.append(Path(image_path))
        self.dimension_calls += 1
        self.known_structure_type = known_structure_type
        return {
            "normalized": {
                "long_width": "4L×2L",
                "single_side": "",
                "dimension_state": "full",
            }
        }


class _FakeCoordinator:
    def __init__(self):
        self.qwen = _FakeQwen()


class FeishuStoreFlowTests(unittest.TestCase):
    def test_8788_store_orientation_is_disabled_by_default_with_opt_in_flag(self):
        defaults = build_parser().parse_args([])
        disabled = build_parser().parse_args(["--disable-store-text-orientation"])
        enabled = build_parser().parse_args(["--enable-store-text-orientation"])

        self.assertFalse(defaults.enable_store_text_orientation)
        self.assertFalse(disabled.enable_store_text_orientation)
        self.assertTrue(enabled.enable_store_text_orientation)

    def test_question_orientation_precedes_classification_and_becomes_store_source(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "question.jpg"
            upright = root / "question.a3-upright.jpg"
            Image.new("RGB", (20, 10), "red").save(source)
            Image.new("RGB", (10, 20), "blue").save(upright)
            orientation_calls = []

            def orient(path):
                orientation_calls.append(Path(path))
                return upright

            coordinator = _FakeCoordinator()
            draft = FeishuStoreService(
                root=root / "main",
                symbolic=root / "symbolic",
                dry_run=True,
                question_orienter=orient,
            ).classify_question(source, coordinator)

        self.assertEqual(orientation_calls, [source.resolve()])
        self.assertEqual(draft.question_image_path, str(upright.resolve()))
        self.assertTrue(coordinator.qwen.image_paths)
        self.assertTrue(
            all(path == upright.resolve() for path in coordinator.qwen.image_paths)
        )

    def test_apply_plan_writes_the_oriented_question_image(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "question.jpg"
            upright = root / "question.a3-upright.jpg"
            answer = root / "answer.jpg"
            Image.new("RGB", (20, 10), "red").save(source)
            Image.new("RGB", (10, 20), "blue").save(upright)
            Image.new("RGB", (8, 8), "white").save(answer)
            service = FeishuStoreService(
                root=root / "main",
                symbolic=root / "symbolic",
                question_orienter=lambda _path: upright,
            )
            draft = service.classify_question(source, _FakeCoordinator())
            draft.answer_image_paths = [str(answer)]

            with patch(
                "scripts.feishu_store_flow.backup_workbook",
                return_value=root / "backup.xlsx",
            ):
                result = service.apply_plan(draft)

            with Image.open(result.plan.question_target) as stored:
                stored_size = stored.size

        self.assertEqual(stored_size, (10, 20))

    def test_arch_store_skips_outer_dimension_recognition(self):
        coordinator = _FakeCoordinator()
        result = classify_store_dimensions(Path("question.jpg"), "拱", coordinator)
        self.assertEqual(result["dimension_state"], "skip")
        self.assertEqual(coordinator.qwen.dimension_calls, 0)

    def test_symbolic_store_classification_carries_structure_and_dimensions(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "main"
            symbolic = Path(temp_dir) / "symbolic"
            draft = FeishuStoreService(root=root, symbolic=symbolic, dry_run=True).classify_question(
                Path(temp_dir) / "question.jpg", _FakeCoordinator()
            )

        self.assertEqual(draft.route, "symbolic")
        self.assertEqual(draft.structure_type, "桁架")
        self.assertEqual(draft.long_width, "4L×2L")
        self.assertEqual(draft.dimension_state, "full")

    def test_append_persists_symbolic_metadata_and_adds_missing_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook = Path(temp_dir) / "2静定结构.xlsx"
            append_excel_record(
                workbook,
                "2静定结构/题目/1.jpg",
                [{"type": "均布", "raw": "q"}],
                structure_type="桁架",
                long_width="4L×2L",
            )
            wb = openpyxl.load_workbook(workbook, read_only=True)
            headers = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
            values = [cell.value for cell in next(wb.active.iter_rows(min_row=2, max_row=2))]
            wb.close()

        row = dict(zip(headers, values))
        self.assertEqual(row["结构类型"], "桁架")
        self.assertEqual(row["长×宽"], "4L×2L")

    def test_confirmation_shows_symbolic_metadata(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            service = FeishuStoreService(root=Path(temp_dir) / "main", symbolic=Path(temp_dir) / "symbolic", dry_run=True)
            draft = service.classify_question(Path(temp_dir) / "question.jpg", _FakeCoordinator())
            draft.answer_image_paths = [str(Path(temp_dir) / "answer.jpg")]
            plan = service.prepare_plan(draft)

        confirmation = format_store_confirmation(plan)
        self.assertIn("结构类型：桁架", confirmation)
        self.assertIn("外围尺寸（长×宽）：4L×2L", confirmation)


if __name__ == "__main__":
    unittest.main()
