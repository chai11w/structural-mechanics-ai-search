"""
Detect unindexed question images and optionally append them to the right bank.

Default mode is dry-run. It scans with audit_unindexed_questions, classifies only
the missing images, routes them to the main or symbolic bank, and writes reports.
Pass --apply to append records to live Excel files after backing them up.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

import search
from multi_agent_pipeline import QwenClassifier, RuleRouter, symbolic_root
from scripts.apply_main_bank_update import normalized_main_loads
from scripts.audit_unindexed_questions import (
    CHAPTERS,
    audit_chapter,
    load_special_index,
)
from scripts.build_symbolic_bank import mapped_symbolic_loads
from scripts.classify_question_bank import normalize_load_item


MAIN_CATEGORIES = {"main_numeric", "main_assigned_symbolic"}


@dataclass
class StorePlan:
    rel_path: str
    chapter: str
    image_path: str
    route: str
    category: str
    target_bank: str
    target_excel: str
    loads: list[dict[str, Any]]
    stored_loads: list[dict[str, Any]]
    chapter_hint: str
    chapter_confidence: float
    chapter_evidence: str
    from_cache: bool
    status: str
    reason: str
    seconds: float
    error: str = ""


@dataclass
class ApplyResult:
    workbook: str
    bank: str
    appended: list[str] = field(default_factory=list)
    skipped_existing: list[str] = field(default_factory=list)
    dry_run: bool = True


def rel_to_chapter(rel_path: str) -> str:
    return rel_path.replace("\\", "/").split("/", 1)[0]


def resolve_image_path(root: Path, rel_path: str) -> Path:
    return root / rel_path.replace("/", "\\")


def loads_for_main(rel_path: str, loads: list[dict[str, Any]], category: str) -> list[dict[str, Any]]:
    record = {"rel_path": rel_path, "loads": loads, "category": category}
    return normalized_main_loads(record)


def loads_for_symbolic(loads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return mapped_symbolic_loads(loads)


def json_loads(loads: list[dict[str, Any]]) -> str:
    return json.dumps({"loads": loads}, ensure_ascii=False)


def sheet_headers(ws) -> dict[str, int]:
    headers = {}
    for index, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = index
    return headers


def ensure_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["题目名称", "荷载"])

    headers = sheet_headers(ws)
    if "题目名称" not in headers or "荷载" not in headers:
        raise ValueError(f"{path}: missing required columns 题目名称/荷载")
    return wb, ws, headers


def existing_rels(ws, question_col: int) -> set[str]:
    values: set[str] = set()
    for row_index in range(2, ws.max_row + 1):
        raw = ws.cell(row_index, question_col).value
        if raw:
            values.add(str(raw).replace("\\", "/").strip())
    return values


def backup_workbook(path: Path, backup_dir: Path, bank: str) -> Path | None:
    if not path.exists():
        return None
    target_dir = backup_dir / bank
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / path.name
    if not target.exists():
        shutil.copy2(path, target)
    return target


def append_plans_to_workbook(
    workbook: Path,
    plans: list[StorePlan],
    *,
    bank: str,
    dry_run: bool,
    backup_dir: Path,
) -> ApplyResult:
    result = ApplyResult(workbook=str(workbook), bank=bank, dry_run=dry_run)
    if not plans:
        return result

    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb, ws, headers = ensure_workbook(workbook)
    question_col = headers["题目名称"]
    loads_col = headers["荷载"]
    known = existing_rels(ws, question_col)

    for plan in plans:
        rel = plan.rel_path.replace("\\", "/")
        if rel in known:
            result.skipped_existing.append(rel)
            continue
        result.appended.append(rel)
        if not dry_run:
            row = [None] * ws.max_column
            row[question_col - 1] = rel
            row[loads_col - 1] = json_loads(plan.stored_loads)
            ws.append(row)
            known.add(rel)

    if not dry_run and result.appended:
        backup_workbook(workbook, backup_dir, bank)
        wb.save(workbook)
    wb.close()
    return result


def collect_missing(root: Path, symbolic: Path, chapters: list[str], special_keys: set[str]) -> list[str]:
    missing: list[str] = []
    for chapter in chapters:
        audit = audit_chapter(root, symbolic, chapter, special_keys)
        missing.extend(audit.missing)
    return sorted(missing, key=lambda item: item.casefold())


def classify_missing(
    rel_paths: list[str],
    *,
    root: Path,
    symbolic: Path,
    qwen: QwenClassifier,
    router: RuleRouter,
    sleep: float,
) -> list[StorePlan]:
    plans: list[StorePlan] = []

    for index, rel_path in enumerate(rel_paths, 1):
        started = time.time()
        chapter = rel_to_chapter(rel_path)
        image_path = resolve_image_path(root, rel_path)
        status = "needs_review"
        reason = ""
        route_name = "needs_review"
        category = "unknown"
        target_bank = ""
        target_excel = ""
        loads: list[dict[str, Any]] = []
        stored_loads: list[dict[str, Any]] = []
        chapter_hint = ""
        chapter_confidence = 0.0
        chapter_evidence = ""
        from_cache = False
        error = ""

        try:
            classified = qwen.classify_image(image_path)
            loads = [
                normalize_load_item(item)
                for item in search.normalize_query_loads(classified.get("loads", []))
                if isinstance(item, dict)
            ]
            route, _load_details = router.route(loads)
            route_name = route.route
            category = route.category
            chapter_hint = str(classified.get("chapter_hint") or "")
            chapter_confidence = float(classified.get("chapter_confidence") or 0.0)
            chapter_evidence = str(classified.get("chapter_evidence") or "")
            from_cache = bool(classified.get("from_cache"))

            if not loads:
                status = "needs_review"
                reason = "未识别到荷载"
            elif chapter not in CHAPTERS:
                status = "needs_review"
                reason = f"章节不在范围内: {chapter}"
            elif route.route == "main" and category in MAIN_CATEGORIES:
                target_bank = "main"
                target_excel = str(root / f"{chapter}.xlsx")
                stored_loads = loads_for_main(rel_path, loads, category)
                status = "ready"
                reason = route.reason
            elif route.route == "symbolic" and category == "symbolic_unassigned":
                target_bank = "symbolic"
                target_excel = str(symbolic / f"{chapter}.xlsx")
                stored_loads = loads_for_symbolic(loads)
                status = "ready"
                reason = route.reason
            else:
                status = "needs_review"
                reason = route.reason
        except Exception as exc:  # noqa: BLE001 - keep batch storage jobs inspectable.
            error = str(exc)
            reason = error

        plan = StorePlan(
            rel_path=rel_path,
            chapter=chapter,
            image_path=str(image_path),
            route=route_name,
            category=category,
            target_bank=target_bank,
            target_excel=target_excel,
            loads=loads,
            stored_loads=stored_loads,
            chapter_hint=chapter_hint,
            chapter_confidence=chapter_confidence,
            chapter_evidence=chapter_evidence,
            from_cache=from_cache,
            status=status,
            reason=reason,
            seconds=round(time.time() - started, 2),
            error=error,
        )
        plans.append(plan)
        print(
            f"[{index}/{len(rel_paths)}] {plan.status} | {plan.target_bank or '-'} | "
            f"{plan.category} | {plan.rel_path} | {plan.seconds}s"
        )
        if sleep:
            time.sleep(sleep)

    return plans


def plan_to_dict(plan: StorePlan) -> dict[str, Any]:
    return {
        "rel_path": plan.rel_path,
        "chapter": plan.chapter,
        "image_path": plan.image_path,
        "route": plan.route,
        "category": plan.category,
        "target_bank": plan.target_bank,
        "target_excel": plan.target_excel,
        "loads": plan.loads,
        "stored_loads": plan.stored_loads,
        "chapter_hint": plan.chapter_hint,
        "chapter_confidence": plan.chapter_confidence,
        "chapter_evidence": plan.chapter_evidence,
        "from_cache": plan.from_cache,
        "status": plan.status,
        "reason": plan.reason,
        "seconds": plan.seconds,
        "error": plan.error,
    }


def apply_ready_plans(
    plans: list[StorePlan],
    *,
    dry_run: bool,
    backup_dir: Path,
) -> list[ApplyResult]:
    grouped: dict[tuple[str, str], list[StorePlan]] = defaultdict(list)
    for plan in plans:
        if plan.status != "ready" or not plan.target_excel or not plan.target_bank:
            continue
        grouped[(plan.target_bank, plan.target_excel)].append(plan)

    results: list[ApplyResult] = []
    for (bank, workbook), items in sorted(grouped.items()):
        results.append(
            append_plans_to_workbook(
                Path(workbook),
                items,
                bank=bank,
                dry_run=dry_run,
                backup_dir=backup_dir,
            )
        )
    return results


def write_reports(
    plans: list[StorePlan],
    apply_results: list[ApplyResult],
    *,
    output_dir: Path,
    root: Path,
    symbolic: Path,
    dry_run: bool,
    special_index: Path,
    backup_dir: Path,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        "dry_run": dry_run,
        "root": str(root),
        "symbolic_root": str(symbolic),
        "special_index": str(special_index),
        "backup_dir": str(backup_dir) if not dry_run else "",
        "total": len(plans),
        "status_counts": dict(Counter(plan.status for plan in plans)),
        "target_counts": dict(Counter(plan.target_bank or "none" for plan in plans)),
        "apply_results": [
            {
                "workbook": result.workbook,
                "bank": result.bank,
                "dry_run": result.dry_run,
                "append_count": len(result.appended),
                "skipped_existing_count": len(result.skipped_existing),
                "appended": result.appended,
                "skipped_existing": result.skipped_existing,
            }
            for result in apply_results
        ],
        "plans": [plan_to_dict(plan) for plan in plans],
    }
    (output_dir / "store_unindexed_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    lines = ["# 漏存题目自动补库报告", ""]
    lines.append(f"- dry_run: {dry_run}")
    lines.append(f"- root: `{root}`")
    lines.append(f"- symbolic_root: `{symbolic}`")
    lines.append(f"- special_index: `{special_index}`")
    if not dry_run:
        lines.append(f"- backup_dir: `{backup_dir}`")
    lines.append(f"- total_missing_processed: {len(plans)}")
    lines.append(f"- status_counts: {dict(Counter(plan.status for plan in plans))}")
    lines.append(f"- target_counts: {dict(Counter(plan.target_bank or 'none' for plan in plans))}")
    lines.append("")

    lines.append("## Apply Results")
    lines.append("")
    if apply_results:
        for result in apply_results:
            lines.append(f"- `{result.workbook}` ({result.bank}) append={len(result.appended)} skipped_existing={len(result.skipped_existing)}")
    else:
        lines.append("- no ready records")
    lines.append("")

    if plans:
        lines.append("## Plans")
        lines.append("")
        for plan in plans:
            lines.append(f"### {plan.rel_path}")
            lines.append("")
            lines.append(f"- status: {plan.status}")
            lines.append(f"- category: {plan.category}")
            lines.append(f"- target: {plan.target_bank or '-'}")
            lines.append(f"- reason: {plan.reason}")
            lines.append(f"- loads: `{json.dumps({'loads': plan.loads}, ensure_ascii=False)}`")
            if plan.stored_loads:
                lines.append(f"- stored_loads: `{json.dumps({'loads': plan.stored_loads}, ensure_ascii=False)}`")
            if plan.error:
                lines.append(f"- error: `{plan.error}`")
            lines.append("")

    (output_dir / "store_unindexed_report.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Detect unindexed question images and append recognized records to main/symbolic Excel.")
    parser.add_argument("--root", default=str(search.ROOT), help="main question bank root")
    parser.add_argument("--symbolic-root", help="symbolic bank root; defaults to <root>_字母库 beside main root")
    parser.add_argument("--chapter", action="append", choices=CHAPTERS, help="chapter to scan; may be repeated")
    parser.add_argument("--special-index", default=str(BASE / "special_unindexed_questions.json"))
    parser.add_argument("--no-special-index", action="store_true", help="ignore special exclusions and process all missing images")
    parser.add_argument("--limit", type=int, default=0, help="process at most N missing images; 0 means all")
    parser.add_argument("--sleep", type=float, default=0.0, help="sleep seconds between Qwen calls")
    parser.add_argument("--timeout", type=int, default=180)
    parser.add_argument("--model", default=None, help="override Qwen model")
    parser.add_argument("--endpoint", default=None, help="override Qwen endpoint")
    parser.add_argument("--no-cache", action="store_true", help="disable Qwen classifier cache")
    parser.add_argument("--output-dir", help="report output directory")
    parser.add_argument("--apply", action="store_true", help="write ready records to Excel; default is dry-run")
    args = parser.parse_args()

    root = Path(args.root)
    symbolic = Path(args.symbolic_root) if args.symbolic_root else symbolic_root(root)
    chapters = args.chapter or CHAPTERS
    special_index = Path(args.special_index)
    special_keys = set() if args.no_special_index else load_special_index(special_index, main_root=root)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir) if args.output_dir else BASE / ".tmp_audit" / f"store_unindexed_{stamp}"
    backup_dir = BASE / "backups" / f"store_unindexed_{stamp}"

    missing = collect_missing(root, symbolic, chapters, special_keys)
    if args.limit:
        missing = missing[: args.limit]

    print(f"root={root}")
    print(f"symbolic_root={symbolic}")
    print(f"missing_to_process={len(missing)}")
    print(f"mode={'apply' if args.apply else 'dry-run'}")

    qwen_kwargs: dict[str, Any] = {
        "timeout": args.timeout,
        "use_cache": not args.no_cache,
    }
    if args.model:
        qwen_kwargs["model"] = args.model
    if args.endpoint:
        qwen_kwargs["endpoint"] = args.endpoint
    qwen = QwenClassifier(**qwen_kwargs)
    router = RuleRouter()

    plans = classify_missing(
        missing,
        root=root,
        symbolic=symbolic,
        qwen=qwen,
        router=router,
        sleep=args.sleep,
    ) if missing else []
    apply_results = apply_ready_plans(plans, dry_run=not args.apply, backup_dir=backup_dir)
    write_reports(
        plans,
        apply_results,
        output_dir=output_dir,
        root=root,
        symbolic=symbolic,
        dry_run=not args.apply,
        special_index=special_index,
        backup_dir=backup_dir,
    )

    ready = sum(1 for plan in plans if plan.status == "ready")
    review = sum(1 for plan in plans if plan.status != "ready")
    appended = sum(len(result.appended) for result in apply_results)
    print(f"ready={ready} needs_review={review} append_candidates={appended}")
    print(f"report={output_dir / 'store_unindexed_report.md'}")
    print(f"json={output_dir / 'store_unindexed_summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
