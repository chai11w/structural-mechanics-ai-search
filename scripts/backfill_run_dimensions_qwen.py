"""Run qwen v4 dimension recognition on the full letter bank, no HTML review.

Builds a versioned manifest from the letter-bank xlsx files (row path + bank
structure type), calls qwen with the current ``DIMENSION_PROMPT`` concurrently
with per-image retries, and writes two artifacts under an ignored output dir:

  - ``results.json`` — per-image normalized rows (model + code sums, verified
    flag, long_width). No review images and no ``review.html`` are produced.
  - ``qwen_v4_backfill_verdicts.json`` — ``{path, long_width}`` rows ready for
    ``scripts/backfill_letter_bank_dimensions.py``.

Verdict rule: every row with a readable ``long_width`` is included — the value
is the code-summed segments when they parse (authoritative) and falls back to
the model total otherwise — while the 10 already human-verified paths are
excluded so model output never overwrites human verdicts. Rows with no readable
long_width (``unknown``, unreadable segments, arches with a null height) stay
blank in the bank, which is the recall-preserving choice for the hard filter.

The API key is read from ``DASHSCOPE_API_KEY`` only; never from a file.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

BASE = Path(__file__).resolve().parent.parent
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))

from scripts.classify_question_bank import DEFAULT_ENDPOINT, DEFAULT_MODEL  # noqa: E402
from scripts.evaluate_structure_dimensions import (  # noqa: E402
    Sample,
    call_qwen,
    load_manifest,
)

DEFAULT_BANK_ROOT = Path("D:/桌面/答疑、帮做/结构力学/帮做_字母库")
DEFAULT_IMAGES_ROOT = Path("D:/桌面/答疑、帮做/结构力学/帮做")
DEFAULT_MANIFEST = BASE / "experiments" / "structure_dimension_eval" / "bank_all_280.json"
DEFAULT_HUMAN_VERDICTS = (
    BASE / "experiments" / "structure_dimension_eval" / "human_verdicts.json"
)
DEFAULT_WORKERS = 10
MAX_ATTEMPTS = 3
RETRY_DELAY_SECONDS = 2
VERDICTS_FILENAME = "qwen_v4_backfill_verdicts.json"


def _sample_id_from_path(path: str) -> str:
    # Sanitize the full path (minus extension) so every unique bank path maps to a
    # unique, human-readable id, e.g. "2\u9759\u5b9a\u7ed3\u6784/1\u5355\u8de8\u6881/\u9898\u76eea/1\u5747/10.jpg" ->
    # "2\u9759\u5b9a\u7ed3\u6784_1\u5355\u8de8\u6881_\u9898\u76eea_1\u5747_10".
    without_ext = path.rsplit(".", 1)[0]
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", without_ext).strip("_")


def build_manifest_from_bank(bank_root: Path, images_root: Path) -> list[dict[str, str]]:
    """Read the letter-bank xlsx files into manifest sample dicts (stable order)."""

    samples: list[dict[str, str]] = []
    seen: set[str] = set()
    for xlsx in sorted(bank_root.glob("*.xlsx")):
        import openpyxl

        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb.worksheets[0]
        for row in range(2, ws.max_row + 1):
            raw_path = ws.cell(row=row, column=1).value
            if not raw_path:
                continue
            path = str(raw_path).replace("\\", "/").strip()
            if not path or path in seen:
                raise ValueError(f"duplicate or empty bank path in {xlsx.name}: {path!r}")
            seen.add(path)
            structure_type = str(ws.cell(row=row, column=3).value or "unknown").strip()
            samples.append(
                {
                    "id": _sample_id_from_path(path),
                    "expected_structure_type": structure_type,
                    "path": path,
                    "selection_note": f"字母库 {xlsx.name} 第 {row} 行",
                }
            )
        wb.close()
    return samples


def write_manifest(path: Path, samples: list[dict[str, str]]) -> None:
    path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "description": (
                    "Full letter-bank dimension-recognition manifest (2026-08-05): all 280 "
                    "letter-bank rows (path + bank structure type). Generated from the bank "
                    "xlsx files; images resolve under the question-bank root."
                ),
                "samples": samples,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def load_human_paths(path: Path) -> set[str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return {
        str(item.get("path") or "").replace("\\", "/").strip()
        for item in data.get("verdicts", [])
        if item.get("path")
    }


def run_one(
    sample: Sample,
    *,
    root: Path,
    api_key: str,
    endpoint: str,
    model: str,
    timeout: int,
) -> dict[str, Any]:
    image_path = root / sample.relative_path
    started = time.perf_counter()
    last_error = ""
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            normalized, usage, raw_content = call_qwen(
                image_path,
                api_key=api_key,
                endpoint=endpoint,
                model=model,
                timeout=timeout,
            )
            return {
                "path": sample.relative_path,
                "expected_structure_type": sample.expected_structure_type,
                "normalized": normalized,
                "usage": usage,
                "raw_content": raw_content,
                "seconds": round(time.perf_counter() - started, 3),
                "attempts": attempt,
            }
        except Exception as exc:  # noqa: BLE001 - retry transient failures, keep the last one
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt < MAX_ATTEMPTS:
                time.sleep(RETRY_DELAY_SECONDS * attempt)
    return {
        "path": sample.relative_path,
        "expected_structure_type": sample.expected_structure_type,
        "error": last_error,
        "seconds": round(time.perf_counter() - started, 3),
        "attempts": MAX_ATTEMPTS,
    }


def build_verdicts(
    results: list[dict[str, Any]], human_paths: set[str]
) -> tuple[list[dict[str, str]], dict[str, list[str]]]:
    """Qwen backfill verdicts: any readable long_width, never over human verdicts.

    A readable value is the code-summed segments when they parse (authoritative)
    and falls back to the model total otherwise; ``dimensions_verified`` is a
    cross-check flag, not a gate on the value, so a correct code sum is written
    even when the model's own estimate disagreed.
    """

    verdicts: list[dict[str, str]] = []
    stats: dict[str, list[str]] = {
        "failed": [],
        "unverified": [],
        "blank": [],
        "human_preserved": [],
    }
    for row in results:
        path = row["path"]
        if "error" in row:
            stats["failed"].append(path)
            continue
        normalized = row.get("normalized") or {}
        long_width = str(normalized.get("long_width") or "").strip()
        if not long_width or long_width == "unknown":
            stats["blank"].append(path)
            continue
        if not normalized.get("dimensions_verified"):
            stats["unverified"].append(path)
        if path in human_paths:
            stats["human_preserved"].append(path)
            continue
        verdicts.append({"path": path, "long_width": long_width})
    verdicts.sort(key=lambda item: item["path"])
    return verdicts, stats


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Run qwen v4 dimension recognition on the whole letter bank."
    )
    parser.add_argument("--bank-root", type=Path, default=DEFAULT_BANK_ROOT)
    parser.add_argument("--images-root", type=Path, default=DEFAULT_IMAGES_ROOT)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="ignored dir for results.json and verdicts (default: .tmp_structure_dimension_eval/backfill_qwen_v4_<date>)",
    )
    parser.add_argument("--qwen-model", default=DEFAULT_MODEL)
    parser.add_argument("--qwen-endpoint", default=DEFAULT_ENDPOINT)
    parser.add_argument("--timeout", type=int, default=120)
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    parser.add_argument(
        "--reuse-results",
        type=Path,
        default=None,
        help="recompute the verdicts from a prior run's results.json without calling qwen",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="smoke-test cap: run only the first N samples (0 = all)",
    )
    args = parser.parse_args()

    bank_root = args.bank_root.resolve()
    images_root = args.images_root.resolve()
    if not bank_root.is_dir():
        raise SystemExit(f"letter bank root missing: {bank_root}")
    if not images_root.is_dir():
        raise SystemExit(f"question image root missing: {images_root}")

    manifest = args.manifest.resolve()
    if args.reuse_results is not None:
        payload = json.loads(args.reuse_results.read_text(encoding="utf-8"))
        results = list(payload.get("results") or [])
        output_dir = args.reuse_results.resolve().parent
        images_root = Path(str(payload.get("images_root") or images_root))
        manifest = Path(str(payload.get("manifest") or manifest))
        qwen_model = str(payload.get("qwen_model") or args.qwen_model)
        print(f"reused_results={args.reuse_results} rows={len(results)}")
    else:
        api_key = os.environ.get("DASHSCOPE_API_KEY", "")
        if not api_key:
            raise SystemExit("DASHSCOPE_API_KEY missing; refusing to read a key from configuration files")
        if manifest.is_file():
            samples = load_manifest(manifest, images_root)
        else:
            sample_dicts = build_manifest_from_bank(bank_root, images_root)
            write_manifest(manifest, sample_dicts)
            samples = load_manifest(manifest, images_root)
            print(f"wrote_manifest={manifest} samples={len(samples)}")
        if args.limit:
            samples = samples[: args.limit]

        output_dir = args.output_dir or (
            Path(__file__).resolve().parent.parent
            / ".tmp_structure_dimension_eval"
            / f"backfill_qwen_v4_{time.strftime('%Y%m%d_%H%M%S')}"
        )
        output_dir.mkdir(parents=True, exist_ok=True)

        results: list[dict[str, Any]] = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {
                pool.submit(
                    run_one,
                    sample,
                    root=images_root,
                    api_key=api_key,
                    endpoint=args.qwen_endpoint,
                    model=args.qwen_model,
                    timeout=args.timeout,
                ): sample.sample_id
                for sample in samples
            }
            for future in concurrent.futures.as_completed(futures):
                row = future.result()
                results.append(row)
                if "error" in row:
                    print(f"qwen failed {row['path']}: {row['error'][:160]}")
                else:
                    print(f"qwen ok {row['path']}")
    results.sort(key=lambda row: row["path"])

    verdicts, stats = build_verdicts(results, load_human_paths(DEFAULT_HUMAN_VERDICTS))
    verified = sum(1 for row in results if (row.get("normalized") or {}).get("dimensions_verified"))

    payload = {
        "schema_version": 1,
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "manifest": str(manifest),
        "images_root": str(images_root),
        "qwen_model": qwen_model,
        "prompt_version": "structure-total-span-height-long-width-v4",
        "workers": args.workers,
        "summary": {
            "total": len(results),
            "ok": sum(1 for row in results if "error" not in row),
            "failed": len(stats["failed"]),
            "verified": verified,
            "verdicts": len(verdicts),
            "blank": len(stats["blank"]),
            "unverified": len(stats["unverified"]),
            "human_preserved": len(stats["human_preserved"]),
        },
        "results": results,
    }
    (output_dir / "results.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / VERDICTS_FILENAME).write_text(
        json.dumps(
            {
                "schema_version": 1,
                "description": (
                    "Qwen v4 dimension backfill verdicts: any readable long_width (code-summed "
                    "segments authoritative, model total fallback), the 10 human-verified rows "
                    "excluded. Ready for scripts/backfill_letter_bank_dimensions.py."
                ),
                "verdicts": verdicts,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"results={output_dir / 'results.json'}")
    print(f"verdicts={output_dir / VERDICTS_FILENAME}")
    print(
        "summary: "
        + json.dumps(payload["summary"], ensure_ascii=False)
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
