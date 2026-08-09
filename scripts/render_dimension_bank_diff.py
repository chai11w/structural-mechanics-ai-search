"""Render a read-only HTML review of current letter-bank vs Qwen dimensions."""

from __future__ import annotations

import argparse
import html
import json
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl


UNKNOWN = "unknown"


def canonical(value: Any) -> str:
    text = str(value or "").strip()
    return text if text else UNKNOWN


def read_bank(root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for book in sorted(root.glob("*.xlsx")):
        # Some valid xlsx writers omit the optional worksheet dimension hint;
        # openpyxl read-only mode then exposes max_row/max_column as None.
        # Normal mode derives the used range from cells and remains read-only in
        # practice because this report never saves the workbook.
        workbook = openpyxl.load_workbook(book, read_only=False, data_only=True)
        sheet = workbook.worksheets[0]
        headers = [canonical(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
        if "长×宽" not in headers:
            workbook.close()
            raise ValueError(f"missing 长×宽 column: {book}")
        dimension_col = headers.index("长×宽") + 1
        for row_number in range(2, sheet.max_row + 1):
            image_path = str(sheet.cell(row_number, 1).value or "").replace("\\", "/").strip()
            if not image_path:
                continue
            rows.append(
                {
                    "path": image_path,
                    "structure_type": canonical(sheet.cell(row_number, 3).value),
                    "current": canonical(sheet.cell(row_number, dimension_col).value),
                    "workbook": book.name,
                    "row": row_number,
                }
            )
        workbook.close()
    return rows


def load_qwen(path: Path) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    results = payload.get("results") or []
    by_path = {str(item.get("path") or "").replace("\\", "/"): item for item in results}
    return by_path, payload


def classify(current: str, proposed: str) -> str:
    if current != UNKNOWN and proposed == UNKNOWN:
        return "现库有值 → 新结果 unknown"
    if current == UNKNOWN and proposed != UNKNOWN:
        return "现库空白 → 新结果有值"
    return "双方有值但不同"


def build_differences(
    bank_rows: list[dict[str, Any]],
    qwen: dict[str, dict[str, Any]],
    *,
    require_all: bool = True,
) -> list[dict[str, Any]]:
    differences: list[dict[str, Any]] = []
    missing = []
    for bank in bank_rows:
        result = qwen.get(bank["path"])
        if result is None:
            missing.append(bank["path"])
            continue
        normalized = result.get("normalized") or {}
        proposed = canonical(normalized.get("long_width"))
        if proposed == bank["current"]:
            continue
        differences.append(
            {
                **bank,
                "proposed": proposed,
                "category": classify(bank["current"], proposed),
                "normalized": normalized,
                "seconds": float(result.get("seconds") or 0),
                "attempts": int(result.get("attempts") or 0),
            }
        )
    if missing and require_all:
        raise ValueError(f"Qwen results missing {len(missing)} bank paths; first={missing[0]}")
    return differences


def _fmt(value: Any) -> str:
    if value is None or value == []:
        return "—"
    if isinstance(value, list):
        return " + ".join(str(item) for item in value) or "—"
    return str(value)


def write_report(
    differences: list[dict[str, Any]], *, total: int, payload: dict[str, Any], images_root: Path, output: Path
) -> dict[str, Any]:
    image_dir = output.parent / "images"
    image_dir.mkdir(parents=True, exist_ok=True)
    category_counts = Counter(row["category"] for row in differences)
    type_counts = Counter(row["structure_type"] for row in differences)
    model_name = str(payload.get("qwen_model") or "Qwen")
    cards: list[str] = []
    conflicts = 0
    for index, row in enumerate(differences, 1):
        normalized = row["normalized"]
        conflict = normalized.get("span_consistent") is False or normalized.get("height_consistent") is False
        conflicts += int(conflict)
        source = images_root / row["path"]
        if not source.is_file():
            raise FileNotFoundError(source)
        suffix = source.suffix.lower() or ".jpg"
        target_name = f"{index:03d}{suffix}"
        shutil.copy2(source, image_dir / target_name)
        status = "尺寸链冲突，需重点复核" if conflict else ("已通过代码校验" if normalized.get("dimensions_verified") else "未完整校验")
        cards.append(
            f"""
<article class="card {('danger' if row['category'].endswith('unknown') else '')}">
  <header><div><span class="number">#{index:03d}</span><span class="tag">{html.escape(row['category'])}</span></div><span>{row['seconds']:.3f}s · {row['attempts']}次</span></header>
  <p class="path">{html.escape(row['path'])}</p>
  <div class="content">
    <figure><img src="images/{target_name}" alt="{html.escape(row['path'])}"></figure>
    <div class="results">
      <section class="current"><h2>题库现有</h2><strong>{html.escape(row['current'])}</strong><p>{html.escape(row['workbook'])} · 第 {row['row']} 行 · {html.escape(row['structure_type'])}</p></section>
      <section class="proposed"><h2>{html.escape(model_name)} 新识别</h2><strong>{html.escape(row['proposed'])}</strong><p>水平总长：{html.escape(_fmt(normalized.get('total_span')))}</p><p>竖直总高：{html.escape(_fmt(normalized.get('total_height')))}</p><p>水平分段：{html.escape(_fmt(normalized.get('horizontal_segments')))}</p><p>竖直分段：{html.escape(_fmt(normalized.get('vertical_segments')))}</p><p class="status {('conflict' if conflict else '')}">{html.escape(status)}</p><p>说明：{html.escape(_fmt(normalized.get('reason')))}</p></section>
    </div>
  </div>
</article>"""
        )

    summary = {
        "total": total,
        "same": total - len(differences),
        "differences": len(differences),
        "categories": dict(category_counts),
        "types": dict(type_counts),
        "conflicts": conflicts,
        "model": payload.get("qwen_model"),
        "workers": payload.get("workers"),
    }
    metrics = "".join(
        f'<div class="metric"><span>{html.escape(label)}</span><strong>{value}</strong></div>'
        for label, value in [
            ("全库题数", total),
            ("结果相同", summary["same"]),
            ("结果不同", summary["differences"]),
            ("现库有值→unknown", category_counts["现库有值 → 新结果 unknown"]),
            ("现库空白→有值", category_counts["现库空白 → 新结果有值"]),
            ("双方有值但不同", category_counts["双方有值但不同"]),
            ("尺寸链冲突", conflicts),
            ("并发数", summary["workers"] or "—"),
        ]
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>现有题库 vs {html.escape(model_name)} 尺寸差异</title>
<style>body{{margin:0;background:#eef2f7;color:#172033;font:15px/1.55 system-ui,"Microsoft YaHei",sans-serif}}main{{max-width:1500px;margin:auto;padding:28px}}h1{{margin:0 0 6px}}.intro{{color:#536176}}.summary{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin:22px 0}}.metric,.card{{background:#fff;border:1px solid #d9e1ec;border-radius:14px;box-shadow:0 2px 10px #1720330b}}.metric{{padding:14px}}.metric span{{display:block;color:#667085}}.metric strong{{font-size:27px}}.card{{padding:18px;margin:18px 0}}.card.danger{{border-left:6px solid #f59e0b}}header{{display:flex;justify-content:space-between;gap:12px;align-items:center}}.number{{font-size:20px;font-weight:750;margin-right:10px}}.tag{{padding:4px 9px;border-radius:999px;background:#e8eef8;color:#334155}}.path{{word-break:break-all;color:#64748b}}.content{{display:grid;grid-template-columns:minmax(430px,1.2fr) minmax(460px,1fr);gap:18px}}figure{{margin:0;border:1px solid #e2e8f0;background:#fff}}img{{display:block;width:100%;max-height:760px;object-fit:contain}}.results{{display:grid;gap:12px}}section{{border:1px solid #dbe3ee;border-radius:11px;padding:15px;background:#f8fafc}}section h2{{font-size:16px;margin:0 0 6px}}section strong{{font-size:28px}}section p{{margin:7px 0;word-break:break-word}}.current{{border-color:#f0b45a;background:#fffaf0}}.proposed{{border-color:#73a3ee;background:#f5f9ff}}.status{{font-weight:700;color:#237a48}}.status.conflict{{color:#b42318}}@media(max-width:900px){{main{{padding:14px}}.content{{grid-template-columns:1fr}}header{{align-items:flex-start;flex-direction:column}}}}</style></head><body><main>
<h1>现有题库 vs {html.escape(model_name)} 尺寸差异</h1><p class="intro">本次 {total} 题参与比较；下方只显示两者不一致的题。新识别仅供人工复核，没有写回数据库。单边未标注时不会自行补另一边。</p>
<div class="summary">{metrics}</div>{''.join(cards) if cards else '<p>全部一致。</p>'}
</main></body></html>""",
        encoding="utf-8",
    )
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bank-root", type=Path, required=True)
    parser.add_argument("--images-root", type=Path, required=True)
    parser.add_argument("--qwen-results", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--allow-partial-results",
        action="store_true",
        help="render only paths present in qwen-results instead of requiring all bank rows",
    )
    args = parser.parse_args()
    bank = read_bank(args.bank_root)
    qwen, payload = load_qwen(args.qwen_results)
    differences = build_differences(
        bank, qwen, require_all=not args.allow_partial_results
    )
    report_total = len(qwen) if args.allow_partial_results else len(bank)
    summary = write_report(differences, total=report_total, payload=payload, images_root=args.images_root, output=args.output)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(args.output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
