"""
Read-only audit for question images that are not present in chapter Excel files.

The scanner compares images under the live question folders with the union of
the main bank Excel files and the separate symbolic-bank Excel files. It does
not classify images, call models, or modify any workbook.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

import search
from multi_agent_pipeline import symbolic_root


CHAPTERS = ["2静定结构", "3静定结构位移", "4力法", "5位移法", "6力矩分配", "7矩阵位移", "8影响线"]
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
SKIP_DIR_PREFIXES = (".", "_")
SKIP_DIR_KEYWORDS = ("答案", "__pycache__")
QUESTION_DIR_PREFIX = "题目"
DEFAULT_SPECIAL_INDEX = BASE / "special_unindexed_questions.json"


@dataclass
class IndexedEntry:
    rel_path: str
    source: str
    workbook: str
    sheet: str
    row: int


@dataclass
class ChapterAudit:
    chapter: str
    scanned_images: list[str] = field(default_factory=list)
    main_indexed: dict[str, IndexedEntry] = field(default_factory=dict)
    symbolic_indexed: dict[str, IndexedEntry] = field(default_factory=dict)
    ignored_special: list[str] = field(default_factory=list)
    missing: list[str] = field(default_factory=list)
    duplicate_index_paths: dict[str, list[IndexedEntry]] = field(default_factory=dict)

    @property
    def indexed_keys(self) -> set[str]:
        return set(self.main_indexed) | set(self.symbolic_indexed)


def path_key(rel_path: str) -> str:
    return rel_path.replace("\\", "/").strip().strip("/").casefold()


def display_rel(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.as_posix()


def normalize_index_path(value: object, *, main_root: Path, workbook_root: Path) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw.lower() == "nan":
        return None

    normalized = raw.replace("\\", "/").strip("/")
    candidate = Path(normalized)
    if candidate.is_absolute():
        for root in (main_root, workbook_root):
            try:
                return candidate.relative_to(root).as_posix()
            except ValueError:
                continue
        return candidate.as_posix()

    while normalized.startswith("./"):
        normalized = normalized[2:]
    return normalized


def load_special_index(path: Path, *, main_root: Path) -> set[str]:
    if not path.exists():
        return set()

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"特殊题目清单不是有效 JSON: {path}: {exc}") from exc

    if isinstance(data, dict):
        entries = data.get("paths", [])
    else:
        entries = data

    keys: set[str] = set()
    if not isinstance(entries, list):
        raise ValueError(f"特殊题目清单 paths 必须是数组: {path}")

    for entry in entries:
        if isinstance(entry, str):
            raw_path = entry
        elif isinstance(entry, dict):
            raw_path = entry.get("path") or entry.get("rel_path")
        else:
            continue
        rel = normalize_index_path(raw_path, main_root=main_root, workbook_root=main_root)
        if rel:
            keys.add(path_key(rel))
    return keys


def should_skip_dir(name: str) -> bool:
    if name.startswith(SKIP_DIR_PREFIXES):
        return True
    return any(keyword in name for keyword in SKIP_DIR_KEYWORDS)


def is_question_image(path: Path, chapter_dir: Path) -> bool:
    if path.suffix.lower() not in IMAGE_EXTENSIONS:
        return False
    try:
        parts = path.relative_to(chapter_dir).parts
    except ValueError:
        parts = path.parts
    return any(part.startswith(QUESTION_DIR_PREFIX) for part in parts[:-1])


def scan_question_images(root: Path, chapter: str) -> list[str]:
    chapter_dir = root / chapter
    if not chapter_dir.is_dir():
        return []

    images: list[str] = []
    for current, dirs, files in os.walk(chapter_dir):
        dirs[:] = [name for name in dirs if not should_skip_dir(name)]
        current_path = Path(current)
        for file_name in files:
            path = current_path / file_name
            if not is_question_image(path, chapter_dir):
                continue
            images.append(display_rel(path, root))
    return sorted(images, key=lambda item: item.casefold())


def read_workbook_index(
    workbook: Path,
    *,
    source: str,
    main_root: Path,
    workbook_root: Path,
) -> tuple[dict[str, IndexedEntry], dict[str, list[IndexedEntry]]]:
    if not workbook.exists():
        return {}, {}

    indexed: dict[str, IndexedEntry] = {}
    duplicates: dict[str, list[IndexedEntry]] = defaultdict(list)

    try:
        wb = load_workbook(workbook, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        print(f"WARN 无法读取 {workbook}: {exc}", file=sys.stderr)
        return {}, {}

    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                continue
            header_map = {
                str(value).strip(): index
                for index, value in enumerate(headers)
                if value is not None
            }
            if "题目名称" not in header_map:
                continue
            col = header_map["题目名称"]
            for row_offset, row in enumerate(rows, start=2):
                if col >= len(row):
                    continue
                rel = normalize_index_path(row[col], main_root=main_root, workbook_root=workbook_root)
                if rel is None:
                    continue
                entry = IndexedEntry(
                    rel_path=rel,
                    source=source,
                    workbook=str(workbook),
                    sheet=ws.title,
                    row=row_offset,
                )
                key = path_key(rel)
                if key in indexed:
                    duplicates[key].append(indexed[key])
                    duplicates[key].append(entry)
                    continue
                indexed[key] = entry
    finally:
        wb.close()

    return indexed, dict(duplicates)


def merge_duplicates(*items: dict[str, list[IndexedEntry]]) -> dict[str, list[IndexedEntry]]:
    merged: dict[str, list[IndexedEntry]] = defaultdict(list)
    for item in items:
        for key, entries in item.items():
            seen = {(entry.source, entry.workbook, entry.sheet, entry.row) for entry in merged[key]}
            for entry in entries:
                identity = (entry.source, entry.workbook, entry.sheet, entry.row)
                if identity not in seen:
                    merged[key].append(entry)
                    seen.add(identity)
    return dict(merged)


def audit_chapter(root: Path, symbolic: Path, chapter: str, special_keys: set[str]) -> ChapterAudit:
    scanned = scan_question_images(root, chapter)
    main_indexed, main_dups = read_workbook_index(
        root / f"{chapter}.xlsx",
        source="main",
        main_root=root,
        workbook_root=root,
    )
    symbolic_indexed, symbolic_dups = read_workbook_index(
        symbolic / f"{chapter}.xlsx",
        source="symbolic",
        main_root=root,
        workbook_root=symbolic,
    )

    indexed_keys = set(main_indexed) | set(symbolic_indexed)
    missing: list[str] = []
    ignored_special: list[str] = []
    for rel in scanned:
        key = path_key(rel)
        if key in indexed_keys:
            continue
        if key in special_keys:
            ignored_special.append(rel)
        else:
            missing.append(rel)
    return ChapterAudit(
        chapter=chapter,
        scanned_images=scanned,
        main_indexed=main_indexed,
        symbolic_indexed=symbolic_indexed,
        ignored_special=ignored_special,
        missing=missing,
        duplicate_index_paths=merge_duplicates(main_dups, symbolic_dups),
    )


def entry_to_dict(entry: IndexedEntry) -> dict:
    return {
        "rel_path": entry.rel_path,
        "source": entry.source,
        "workbook": entry.workbook,
        "sheet": entry.sheet,
        "row": entry.row,
    }


def audit_to_dict(audit: ChapterAudit) -> dict:
    return {
        "chapter": audit.chapter,
        "scanned_image_count": len(audit.scanned_images),
        "main_indexed_count": len(audit.main_indexed),
        "symbolic_indexed_count": len(audit.symbolic_indexed),
        "indexed_union_count": len(audit.indexed_keys),
        "ignored_special_count": len(audit.ignored_special),
        "ignored_special": audit.ignored_special,
        "missing_count": len(audit.missing),
        "missing": audit.missing,
        "duplicate_index_paths": {
            key: [entry_to_dict(entry) for entry in entries]
            for key, entries in audit.duplicate_index_paths.items()
        },
    }


def write_report(
    audits: list[ChapterAudit],
    *,
    root: Path,
    symbolic: Path,
    report_path: Path,
    json_path: Path | None,
    limit: int,
    special_index: Path,
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    total_scanned = sum(len(item.scanned_images) for item in audits)
    total_ignored_special = sum(len(item.ignored_special) for item in audits)
    total_missing = sum(len(item.missing) for item in audits)
    total_dups = sum(len(item.duplicate_index_paths) for item in audits)

    lines = [
        "# 漏存题目路径扫描报告",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 主库根目录：`{root}`",
        f"- 字母库目录：`{symbolic}`",
        f"- 特殊题目清单：`{special_index}`",
        f"- 扫描题图：{total_scanned}",
        f"- 特殊排除：{total_ignored_special}",
        f"- 疑似漏存：{total_missing}",
        f"- Excel 重复路径：{total_dups}",
        "",
    ]

    for audit in audits:
        lines.extend([
            f"## {audit.chapter}",
            "",
            f"- 扫描题图：{len(audit.scanned_images)}",
            f"- 主库已索引：{len(audit.main_indexed)}",
            f"- 字母库已索引：{len(audit.symbolic_indexed)}",
            f"- 已索引并集：{len(audit.indexed_keys)}",
            f"- 特殊排除：{len(audit.ignored_special)}",
            f"- 疑似漏存：{len(audit.missing)}",
            "",
        ])
        if audit.ignored_special:
            shown = audit.ignored_special[:limit] if limit > 0 else audit.ignored_special
            lines.append("### 特殊排除路径")
            lines.append("")
            for rel in shown:
                lines.append(f"- `{rel}`")
            if limit > 0 and len(audit.ignored_special) > limit:
                lines.append(f"- ... 还有 {len(audit.ignored_special) - limit} 条未展示")
            lines.append("")
        if audit.missing:
            shown = audit.missing[:limit] if limit > 0 else audit.missing
            lines.append("### 疑似漏存路径")
            lines.append("")
            for rel in shown:
                lines.append(f"- `{rel}`")
            if limit > 0 and len(audit.missing) > limit:
                lines.append(f"- ... 还有 {len(audit.missing) - limit} 条未展示")
            lines.append("")
        if audit.duplicate_index_paths:
            lines.append("### Excel 重复路径")
            lines.append("")
            for key, entries in sorted(audit.duplicate_index_paths.items()):
                lines.append(f"- `{key}`")
                for entry in entries[:5]:
                    lines.append(f"  - {entry.source}: `{entry.workbook}` / {entry.sheet}!{entry.row}")
                if len(entries) > 5:
                    lines.append(f"  - ... 还有 {len(entries) - 5} 处")
            lines.append("")

    report_path.write_text("\n".join(lines), encoding="utf-8")

    if json_path:
        json_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "root": str(root),
            "symbolic_root": str(symbolic),
            "special_index": str(special_index),
            "total_scanned": total_scanned,
            "total_ignored_special": total_ignored_special,
            "total_missing": total_missing,
            "chapters": [audit_to_dict(item) for item in audits],
        }
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def print_summary(audits: Iterable[ChapterAudit], report_path: Path) -> None:
    audits = list(audits)
    total_scanned = sum(len(item.scanned_images) for item in audits)
    total_ignored_special = sum(len(item.ignored_special) for item in audits)
    total_missing = sum(len(item.missing) for item in audits)
    print(f"scanned={total_scanned} special={total_ignored_special} missing={total_missing}")
    for audit in audits:
        print(
            f"{audit.chapter}\t"
            f"scan={len(audit.scanned_images)}\t"
            f"main={len(audit.main_indexed)}\t"
            f"symbolic={len(audit.symbolic_indexed)}\t"
            f"special={len(audit.ignored_special)}\t"
            f"missing={len(audit.missing)}"
        )
    print(f"report={report_path}")


def unique_report_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for index in range(2, 1000):
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成不重复报告路径: {path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan question image paths missing from main/symbolic Excel indexes.")
    parser.add_argument("--root", default=str(search.ROOT), help="main question bank root")
    parser.add_argument("--symbolic-root", help="symbolic bank root; defaults to <root>_字母库 beside main root")
    parser.add_argument("--chapter", action="append", choices=CHAPTERS, help="chapter to scan; may be repeated")
    parser.add_argument("--special-index", default=str(DEFAULT_SPECIAL_INDEX), help="JSON list of intentionally unindexed special question paths")
    parser.add_argument("--no-special-index", action="store_true", help="ignore the special question index and report all unindexed images")
    parser.add_argument("--report", help="markdown report path; default under .tmp_audit")
    parser.add_argument("--json-out", help="optional JSON report path")
    parser.add_argument("--limit", type=int, default=0, help="max missing paths shown per chapter in markdown; 0 means all")
    parser.add_argument("--fail-on-missing", action="store_true", help="exit 1 when any missing path is found")
    args = parser.parse_args()

    root = Path(args.root)
    symbolic = Path(args.symbolic_root) if args.symbolic_root else symbolic_root(root)
    chapters = args.chapter or CHAPTERS
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = Path(args.report) if args.report else unique_report_path(BASE / ".tmp_audit" / f"unindexed_questions_{stamp}.md")
    json_path = Path(args.json_out) if args.json_out else None
    special_index = Path(args.special_index)
    special_keys = set() if args.no_special_index else load_special_index(special_index, main_root=root)

    audits = [audit_chapter(root, symbolic, chapter, special_keys) for chapter in chapters]
    write_report(
        audits,
        root=root,
        symbolic=symbolic,
        report_path=report_path,
        json_path=json_path,
        limit=args.limit,
        special_index=special_index,
    )
    print_summary(audits, report_path)

    if args.fail_on_missing and any(audit.missing for audit in audits):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
