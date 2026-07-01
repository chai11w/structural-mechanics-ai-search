"""Controlled tools for the question-bank agent MVP.

The agent may plan with these tools, but file-changing tools only apply after
an explicit confirmation. Dry-run mode returns the plan without touching the
live bank.
"""

from __future__ import annotations

import json
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import search
from multi_agent_pipeline import symbolic_root
from scripts.feishu_store_flow import write_as_jpeg
from scripts.tiku_agent_memory import AgentOperation, OperationMemory


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}


@dataclass
class QuestionInspection:
    chapter: str
    question_no: int
    question_files: list[Path] = field(default_factory=list)
    answer_files: list[Path] = field(default_factory=list)
    excel_rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class AgentToolResult:
    ok: bool
    action: str
    message: str
    dry_run: bool = False
    details: dict[str, Any] = field(default_factory=dict)


@dataclass
class ReplaceAnswerPlan:
    chapter: str
    question_no: int
    question_refs: list[str]
    new_answer_images: list[Path]
    old_answer_files: list[Path]
    target_answer_files: list[Path]
    backup_dir: Path


@dataclass
class SoftDeletePlan:
    chapter: str
    question_no: int
    question_refs: list[str]
    question_files: list[Path]
    answer_files: list[Path]
    excel_rows: list[dict[str, Any]]
    delete_dir: Path


class TikuAgentTools:
    def __init__(
        self,
        *,
        root: Path | None = None,
        memory: OperationMemory | None = None,
        dry_run: bool = False,
    ) -> None:
        self.root = Path(root or search.ROOT)
        self.symbolic = symbolic_root(self.root)
        self.memory = memory or OperationMemory()
        self.dry_run = dry_run

    def inspect_question(self, chapter: str, question_no: int) -> QuestionInspection:
        chapter_dir = self.root / chapter
        question_files = find_question_files(chapter_dir, question_no)
        answer_files: list[Path] = []
        for question_file in question_files:
            for answer_file in search.find_answer_files(question_file):
                if answer_file not in answer_files:
                    answer_files.append(answer_file)
        excel_rows = find_excel_rows(self.root, self.symbolic, chapter, question_no)
        return QuestionInspection(chapter, question_no, question_files, answer_files, excel_rows)

    def plan_replace_answer(
        self,
        chapter: str,
        question_no: int,
        new_answer_images: list[Path],
    ) -> ReplaceAnswerPlan:
        if not new_answer_images:
            raise ValueError("还没有收到新答案图")
        inspection = self.inspect_question(chapter, question_no)
        question_refs = unique_question_refs(self.root, inspection)
        ensure_single_question_ref(chapter, question_no, question_refs)
        answer_dir = answer_dir_for(chapter, inspection)
        target_answer_files = [
            answer_dir / answer_name(question_no, index)
            for index in range(len(new_answer_images))
        ]
        backup_dir = backup_root("replace_answer") / chapter / str(question_no)
        return ReplaceAnswerPlan(
            chapter=chapter,
            question_no=question_no,
            question_refs=question_refs,
            new_answer_images=[Path(path) for path in new_answer_images],
            old_answer_files=inspection.answer_files,
            target_answer_files=target_answer_files,
            backup_dir=backup_dir,
        )

    def apply_replace_answer(self, plan: ReplaceAnswerPlan, *, user_id: str | None = None) -> AgentToolResult:
        if self.dry_run:
            return AgentToolResult(
                True,
                "replace_answer",
                format_replace_answer_plan(plan, dry_run=True),
                dry_run=True,
                details=replace_plan_details(plan),
            )

        plan.backup_dir.mkdir(parents=True, exist_ok=True)
        for old_file in plan.old_answer_files:
            if old_file.exists():
                shutil.copy2(old_file, plan.backup_dir / old_file.name)
                old_file.unlink()
        for source, target in zip(plan.new_answer_images, plan.target_answer_files):
            target.parent.mkdir(parents=True, exist_ok=True)
            write_as_jpeg(Path(source), target)
        operation = AgentOperation(
            "replace_answer",
            "success",
            chapter=plan.chapter,
            question_no=plan.question_no,
            user_id=user_id,
            backup=str(plan.backup_dir),
            details=replace_plan_details(plan),
        )
        self.memory.append(operation)
        return AgentToolResult(
            True,
            "replace_answer",
            f"已替换 {plan.chapter} 第 {plan.question_no} 题答案，共 {len(plan.target_answer_files)} 张。",
            details=replace_plan_details(plan),
        )

    def plan_soft_delete(self, chapter: str, question_no: int) -> SoftDeletePlan:
        inspection = self.inspect_question(chapter, question_no)
        question_refs = unique_question_refs(self.root, inspection)
        ensure_single_question_ref(chapter, question_no, question_refs)
        return SoftDeletePlan(
            chapter=chapter,
            question_no=question_no,
            question_refs=question_refs,
            question_files=inspection.question_files,
            answer_files=inspection.answer_files,
            excel_rows=inspection.excel_rows,
            delete_dir=deleted_root() / chapter / str(question_no),
        )

    def apply_soft_delete(self, plan: SoftDeletePlan, *, user_id: str | None = None) -> AgentToolResult:
        if self.dry_run:
            return AgentToolResult(
                True,
                "soft_delete_question",
                format_soft_delete_plan(plan, dry_run=True),
                dry_run=True,
                details=soft_delete_plan_details(plan),
            )

        plan.delete_dir.mkdir(parents=True, exist_ok=True)
        moved_files: list[str] = []
        for source in [*plan.question_files, *plan.answer_files]:
            if source.exists():
                target = unique_target(plan.delete_dir / source.name)
                shutil.move(str(source), str(target))
                moved_files.append(str(target))

        backup_dir = backup_root("soft_delete_question")
        backup_dir.mkdir(parents=True, exist_ok=True)
        removed_rows: list[dict[str, Any]] = []
        for workbook in sorted({Path(row["workbook"]) for row in plan.excel_rows}):
            shutil.copy2(workbook, backup_dir / workbook.name)
            removed_rows.extend(remove_excel_rows(workbook, plan.question_refs))

        operation = AgentOperation(
            "soft_delete_question",
            "success",
            chapter=plan.chapter,
            question_no=plan.question_no,
            user_id=user_id,
            backup=str(backup_dir),
            details={**soft_delete_plan_details(plan), "moved_files": moved_files, "removed_rows": removed_rows},
        )
        self.memory.append(operation)
        return AgentToolResult(
            True,
            "soft_delete_question",
            f"已软删除 {plan.chapter} 第 {plan.question_no} 题，文件已移动到 {plan.delete_dir}。",
            details=operation.details,
        )


def find_question_files(chapter_dir: Path, question_no: int) -> list[Path]:
    if not chapter_dir.exists():
        return []
    found: list[Path] = []
    for path in chapter_dir.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_SUFFIXES:
            continue
        if not any(part.startswith("题目") for part in path.parts):
            continue
        if path.stem == str(question_no):
            found.append(path)
    return sorted(found, key=lambda item: item.as_posix())


def find_excel_rows(root: Path, symbolic: Path, chapter: str, question_no: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for bank, workbook in [("main", root / f"{chapter}.xlsx"), ("symbolic", symbolic / f"{chapter}.xlsx")]:
        if not workbook.exists():
            continue
        wb = load_workbook(workbook)
        ws = wb.active
        headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value}
        question_col = headers.get("题目名称")
        loads_col = headers.get("荷载")
        if not question_col:
            wb.close()
            continue
        for row_index in range(2, ws.max_row + 1):
            rel = str(ws.cell(row=row_index, column=question_col).value or "")
            if Path(rel.replace("\\", "/")).stem != str(question_no):
                continue
            rows.append({
                "bank": bank,
                "workbook": str(workbook),
                "row": row_index,
                "question": rel,
                "loads": ws.cell(row=row_index, column=loads_col).value if loads_col else "",
            })
        wb.close()
    return rows


def remove_excel_rows(workbook: Path, question_refs: list[str]) -> list[dict[str, Any]]:
    wb = load_workbook(workbook)
    ws = wb.active
    headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1]) if cell.value}
    question_col = headers.get("题目名称")
    removed: list[dict[str, Any]] = []
    ref_set = {ref.replace("\\", "/") for ref in question_refs}
    if question_col:
        for row_index in range(ws.max_row, 1, -1):
            rel = str(ws.cell(row=row_index, column=question_col).value or "")
            if rel.replace("\\", "/") in ref_set:
                removed.append({"workbook": str(workbook), "row": row_index, "question": rel})
                ws.delete_rows(row_index, 1)
    wb.save(workbook)
    wb.close()
    return list(reversed(removed))


def answer_dir_for(chapter: str, inspection: QuestionInspection) -> Path:
    if inspection.answer_files:
        return inspection.answer_files[0].parent
    if inspection.question_files:
        question = inspection.question_files[0]
        parts = list(question.parts)
        for index in range(len(parts) - 1, -1, -1):
            if parts[index].startswith("题目"):
                return Path(*parts[:index]) / "答案"
    return Path(search.ROOT) / chapter / "答案"


def unique_question_refs(root: Path, inspection: QuestionInspection) -> list[str]:
    refs: set[str] = set()
    for path in inspection.question_files:
        try:
            refs.add(path.relative_to(root).as_posix())
        except ValueError:
            refs.add(str(path))
    for row in inspection.excel_rows:
        question = str(row.get("question") or "").replace("\\", "/")
        if question:
            refs.add(question)
    return sorted(refs)


def ensure_single_question_ref(chapter: str, question_no: int, refs: list[str]) -> None:
    if not refs:
        raise ValueError(f"未找到 {chapter} 第 {question_no} 题")
    if len(refs) > 1:
        preview = "、".join(refs[:5])
        suffix = "..." if len(refs) > 5 else ""
        raise ValueError(
            f"{chapter} 第 {question_no} 题不唯一，匹配到 {len(refs)} 个位置：{preview}{suffix}。"
            "请先用“查看”确认具体路径，第一版 Agent 不会按模糊题号执行删除或替换。"
        )


def answer_name(question_no: int, index: int) -> str:
    return f"{question_no}{'+' * index}.jpg"


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def backup_root(action: str) -> Path:
    return Path(__file__).resolve().parents[1] / "backups" / f"agent_{action}_{timestamp()}"


def deleted_root() -> Path:
    return Path(__file__).resolve().parents[1] / "deleted" / f"agent_{timestamp()}"


def unique_target(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"cannot find unique target for {path}")


def replace_plan_details(plan: ReplaceAnswerPlan) -> dict[str, Any]:
    return {
        "chapter": plan.chapter,
        "question_no": plan.question_no,
        "question_refs": plan.question_refs,
        "old_answer_files": [str(path) for path in plan.old_answer_files],
        "new_answer_images": [str(path) for path in plan.new_answer_images],
        "target_answer_files": [str(path) for path in plan.target_answer_files],
        "backup_dir": str(plan.backup_dir),
    }


def soft_delete_plan_details(plan: SoftDeletePlan) -> dict[str, Any]:
    return {
        "chapter": plan.chapter,
        "question_no": plan.question_no,
        "question_refs": plan.question_refs,
        "question_files": [str(path) for path in plan.question_files],
        "answer_files": [str(path) for path in plan.answer_files],
        "excel_rows": plan.excel_rows,
        "delete_dir": str(plan.delete_dir),
    }


def format_inspection(inspection: QuestionInspection) -> str:
    lines = [
        f"题目信息：{inspection.chapter} 第 {inspection.question_no} 题",
        f"题图：{len(inspection.question_files)} 张",
    ]
    lines.extend(short_path(path) for path in inspection.question_files)
    lines.append(f"答案：{len(inspection.answer_files)} 张")
    lines.extend(short_path(path) for path in inspection.answer_files)
    lines.append(f"Excel：{len(inspection.excel_rows)} 行")
    for row in inspection.excel_rows:
        loads = row.get("loads")
        try:
            loads_text = json.dumps(json.loads(loads), ensure_ascii=False) if isinstance(loads, str) else str(loads)
        except json.JSONDecodeError:
            loads_text = str(loads)
        lines.append(f"{row['bank']} row {row['row']}：{row['question']} 荷载={loads_text}")
    return "\n".join(lines)


def format_replace_answer_plan(plan: ReplaceAnswerPlan, *, dry_run: bool = False) -> str:
    prefix = "[dry-run] " if dry_run else ""
    lines = [
        f"{prefix}准备执行：替换答案",
        "",
        f"章节：{plan.chapter}",
        f"题号：{plan.question_no}",
        "定位：" + "、".join(plan.question_refs),
        f"原答案：{len(plan.old_answer_files)} 张",
        f"新答案：{len(plan.new_answer_images)} 张",
        "目标：",
    ]
    lines.extend(short_path(path) for path in plan.target_answer_files)
    lines.extend(["", "回复：", "1  确认执行", "0  取消"])
    return "\n".join(lines)


def format_soft_delete_plan(plan: SoftDeletePlan, *, dry_run: bool = False) -> str:
    prefix = "[dry-run] " if dry_run else ""
    lines = [
        f"{prefix}准备执行：软删除题目",
        "",
        f"章节：{plan.chapter}",
        f"题号：{plan.question_no}",
        "定位：" + "、".join(plan.question_refs),
        f"题图：{len(plan.question_files)} 张",
        f"答案：{len(plan.answer_files)} 张",
        f"Excel：{len(plan.excel_rows)} 行会移除",
        f"移动到：{plan.delete_dir}",
        "",
        "回复：",
        "1  确认删除",
        "0  取消",
    ]
    return "\n".join(lines)


def short_path(path: Path) -> str:
    try:
        return path.relative_to(search.ROOT).as_posix()
    except ValueError:
        return str(path)
