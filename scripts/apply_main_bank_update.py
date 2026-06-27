"""
Apply the reviewed main-bank update to live Excel indexes.

This removes existing pure unassigned-symbol rows from the main bank and
appends missing numeric / assigned-symbol rows. It does not touch images or
review-only rows.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent.parent
DEFAULT_EXISTING_RESULTS = BASE / ".tmp_symbol_sheets" / "classify_existing_full_qwen_no_thinking" / "classification_results.json"
DEFAULT_MISSING_RESULTS = BASE / ".tmp_symbol_sheets" / "classify_missing_full_qwen_no_thinking" / "classification_results.json"
MAIN_KEEP_CATEGORIES = {"main_numeric", "main_assigned_symbolic"}


def load_config() -> dict:
    cfg: dict = {}
    for name in ("config.json", "config.local.json"):
        path = BASE / name
        if path.exists():
            try:
                cfg.update(json.loads(path.read_text(encoding="utf-8")))
            except json.JSONDecodeError:
                pass
    return cfg


def configured_root() -> Path:
    cfg = load_config()
    return Path(cfg.get("root") or r"D:\桌面\答疑、帮做\结构力学\帮做")


def load_records(path: Path) -> list[dict]:
    return json.loads(path.read_text(encoding="utf-8"))


def _clean_symbol(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text or "").lower())


def _canonical_expr(text: str) -> str:
    value = str(text or "").strip().lower()
    value = value.replace("_", "")
    value = value.replace(" ", "")
    value = value.replace("×", "*").replace("·", "*")
    return re.sub(r"[^a-z0-9./*]", "", value)


def _format_number(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.6g}"


def _normalize_unit(unit: str) -> str:
    unit = unit.replace("KN", "kN").replace("kn", "kN")
    unit = unit.replace("K", "k")
    return unit


LOAD_UNIT_PATTERNS = [
    r"kN[·.\-*×]?m",
    r"N[·.\-*×]?m",
    r"kN/m",
    r"N/m",
    r"kN",
    r"(?<=\d)k\b",
    r"\bN\b",
]


def strip_load_unit(raw: object) -> str:
    text = str(raw or "").strip()
    text = re.sub(r"\s+", "", text)
    text = text.replace("KN", "kN").replace("kn", "kN")
    for pattern in LOAD_UNIT_PATTERNS:
        text = re.sub(pattern, "", text, flags=re.I)
    return text.strip(".·*-×/")


def _assignment_map(loads: list[dict]) -> dict[str, tuple[float, str]]:
    assignments: dict[str, tuple[float, str]] = {}
    pattern = re.compile(
        r"^\s*([A-Za-z][A-Za-z0-9_]*)\s*=\s*(\d+(?:\.\d+)?)\s*"
        r"((?:kN|KN|N)(?:/m|[·.]m|m)?)?\s*$"
    )
    for item in loads:
        raw = str(item.get("raw", "")).strip()
        match = pattern.match(raw)
        if not match:
            continue
        symbol, value, unit = match.groups()
        assignments[_clean_symbol(symbol)] = (float(value), _normalize_unit(unit or ""))
    return assignments


def _expand_assigned_symbol(raw: str, assignments: dict[str, tuple[float, str]]) -> str | None:
    expr = str(raw or "").strip()
    if "=" in expr or not assignments:
        return None
    compact = _canonical_expr(expr.replace("²", "2").replace("^2", "2"))
    if not compact:
        return None

    for symbol in sorted(assignments, key=len, reverse=True):
        value, unit = assignments[symbol]
        if compact == symbol:
            return _format_number(value)
        match = re.fullmatch(rf"(\d+(?:\.\d+)?)?{re.escape(symbol)}(?:/(\d+(?:\.\d+)?))?", compact)
        if match:
            coeff = float(match.group(1) or 1)
            denom = float(match.group(2) or 1)
            return _format_number(value * coeff / denom)
    return None


def normalized_main_loads(record: dict) -> list[dict]:
    loads = [dict(item) for item in record.get("loads", [])]
    assignments = _assignment_map(loads)
    normalized = []
    for item in loads:
        expanded = _expand_assigned_symbol(str(item.get("raw", "")), assignments)
        if expanded:
            item["raw"] = expanded
        else:
            item["raw"] = strip_load_unit(item.get("raw", ""))
        normalized.append(item)
    return normalized


def loads_json(record: dict) -> str:
    return json.dumps({"loads": normalized_main_loads(record)}, ensure_ascii=False)


def sheet_headers(ws) -> dict[str, int]:
    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value)] = idx
    return headers


def ensure_columns(ws) -> tuple[int, int]:
    headers = sheet_headers(ws)
    if "题目名称" not in headers or "荷载" not in headers:
        raise ValueError(f"{ws.title}: missing required columns")
    return headers["题目名称"], headers["荷载"]


def apply_chapter_update(xlsx_path: Path, delete_rels: set[str], add_records: list[dict], dry_run: bool) -> dict:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    question_col, loads_col = ensure_columns(ws)

    deleted_rows = []
    existing_rels = set()
    for row_idx in range(2, ws.max_row + 1):
        rel = str(ws.cell(row_idx, question_col).value or "").replace("\\", "/")
        if rel:
            existing_rels.add(rel)
        if rel in delete_rels:
            deleted_rows.append((row_idx, rel))

    append_records = [
        record for record in add_records
        if record["rel_path"] not in existing_rels and record["rel_path"] not in delete_rels
    ]

    if not dry_run:
        for row_idx, _rel in sorted(deleted_rows, reverse=True):
            ws.delete_rows(row_idx, 1)

        headers = sheet_headers(ws)
        max_col = ws.max_column
        for record in append_records:
            row = [None] * max_col
            row[headers["题目名称"] - 1] = record["rel_path"]
            row[headers["荷载"] - 1] = loads_json(record)
            ws.append(row)

        wb.save(xlsx_path)

    return {
        "file": xlsx_path.name,
        "delete_count": len(deleted_rows),
        "append_count": len(append_records),
        "deleted": [{"row": row_idx, "rel_path": rel} for row_idx, rel in deleted_rows],
        "appended": [
            {"rel_path": r["rel_path"], "category": r["category"], "loads": normalized_main_loads(r)}
            for r in append_records
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply reviewed main-bank Excel update.")
    parser.add_argument("--root", help="live question bank root")
    parser.add_argument("--existing-results", default=str(DEFAULT_EXISTING_RESULTS))
    parser.add_argument("--missing-results", default=str(DEFAULT_MISSING_RESULTS))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--output-dir")
    args = parser.parse_args()

    root = Path(args.root) if args.root else configured_root()
    existing = load_records(Path(args.existing_results))
    missing = load_records(Path(args.missing_results))

    delete_records = [r for r in existing if r.get("category") == "symbolic_unassigned"]
    add_records = [r for r in missing if r.get("category") in MAIN_KEEP_CATEGORIES]

    deletes_by_chapter: dict[str, set[str]] = defaultdict(set)
    adds_by_chapter: dict[str, list[dict]] = defaultdict(list)
    for record in delete_records:
        deletes_by_chapter[record["chapter"]].add(record["rel_path"])
    for record in add_records:
        adds_by_chapter[record["chapter"]].append(record)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir) if args.output_dir else BASE / ".tmp_symbol_sheets" / f"main_update_apply_{timestamp}"
    output_dir.mkdir(parents=True, exist_ok=True)

    chapter_names = sorted(set(deletes_by_chapter) | set(adds_by_chapter))
    results = []
    for chapter in chapter_names:
        xlsx_path = root / f"{chapter}.xlsx"
        if not xlsx_path.exists():
            raise FileNotFoundError(xlsx_path)
        results.append(
            apply_chapter_update(
                xlsx_path,
                deletes_by_chapter.get(chapter, set()),
                adds_by_chapter.get(chapter, []),
                dry_run=args.dry_run,
            )
        )

    summary = {
        "dry_run": args.dry_run,
        "root": str(root),
        "delete_total": sum(r["delete_count"] for r in results),
        "append_total": sum(r["append_count"] for r in results),
        "by_file": results,
    }

    (output_dir / "main_update_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = ["# 主库更新应用报告", ""]
    lines.append(f"- dry_run: {args.dry_run}")
    lines.append(f"- root: `{root}`")
    lines.append(f"- delete_total: {summary['delete_total']}")
    lines.append(f"- append_total: {summary['append_total']}")
    lines.append("")
    for item in results:
        lines.append(f"## {item['file']}")
        lines.append(f"- delete_count: {item['delete_count']}")
        lines.append(f"- append_count: {item['append_count']}")
        if item["deleted"]:
            lines.append("- deleted:")
            for row in item["deleted"]:
                lines.append(f"  - row {row['row']}: `{row['rel_path']}`")
        if item["appended"]:
            lines.append("- appended:")
            for row in item["appended"]:
                lines.append(f"  - `{row['rel_path']}` | {row['category']} | {json.dumps(row['loads'], ensure_ascii=False)}")
        lines.append("")
    (output_dir / "main_update_summary.md").write_text("\n".join(lines), encoding="utf-8")

    print(f"dry_run={args.dry_run}")
    print(f"delete_total={summary['delete_total']}")
    print(f"append_total={summary['append_total']}")
    for item in results:
        print(f"{item['file']}\tdelete={item['delete_count']}\tappend={item['append_count']}")
    print(f"report={output_dir / 'main_update_summary.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
